from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from rest_framework import status, viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from users.models import Profile, Satker

from django.db.models import Count, Q

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import re
import os
import codecs
import shutil
import datetime
import calendar
import pandas as pd
from django.conf import settings

from sidamas import pagination

from kegiatan import models
from kegiatan.api import serializers
from kegiatan.api import filters
from users.serializers import SatkerSerializer
from users.models import Satker

from kegiatan.api.helpers.api_helpers import (
    get_data_list,
    get_data_list_queryset,
    get_data_list_bnnk,
    get_filtered_data,
    set_created_kegiatan_status,
    get_tanggal_kegiatan,
    
    kirim_kegiatan_helper,
    delete_all_kegiatan_helper,
    aksi_semua_kegiatan,
)

# ======= BINAAN TEKNIS =======
class DAYATIF_BINAAN_TEKNIS_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_BINAAN_TEKNIS.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_BINAAN_TEKNIS_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.DAYATIF_BINAAN_TEKNIS_Filters
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_BINAAN_TEKNIS_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_BINAAN_TEKNIS, serializers.DAYATIF_BINAAN_TEKNIS_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_BINAAN_TEKNIS, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_BINAAN_TEKNIS)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_BINAAN_TEKNIS, request)
            
    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_BINAAN_TEKNIS.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'jumlah_hari_pelaksanaan', 'satker_target', 'satker_target__nama_satker', 'jumlah_peserta', 'tujuan', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')
        
        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'jumlah_hari_pelaksanaan': item['jumlah_hari_pelaksanaan'],
                'satker_target': item['satker_target'],
                'nama_satker_target': item['satker_target__nama_satker'],
                'jumlah_peserta': item['jumlah_peserta'],
                'tujuan': item['tujuan'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data
                            
    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        satker = self.request.user.profile.satker
        filter_status_pengiriman = request.data.get("status", None)
        filter_rentang_waktu = request.data.get("waktu", None)
        
        serialized_data = self.get_flat_values(request, filter_status_pengiriman, filter_rentang_waktu)
        
        # return Response({
        #     'status': True,
        #     'status': filter_status_pengiriman,
        #     'rentang_waktu': filter_rentang_waktu,
        #     'data': len(serialized_data),
        # })
        
        # ======= GET FILE NAME =======
        file_name = ''
                    
        tahun = datetime.datetime.now().year
        base_path = 'media/kegiatan/binaan_teknis/exported'
        main_name = 'REKAPITULASI PEMBINAAN TEKNIS'
        
        if satker.level != 2:
            if filter_rentang_waktu == 'semua':
                file_name = f'{main_name} {satker.nama_satker.upper()} TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan1':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 1 (Januari - Maret) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan2':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 2 (April - Juni) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan3':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 3 (July - September) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan4':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 4 (Oktober - Desember) TAHUN {tahun}'
            elif filter_rentang_waktu == 'hari_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%d %B %Y")}'
            elif filter_rentang_waktu == 'minggu_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%A, %d %B %Y")}'
            elif filter_rentang_waktu == 'bulan_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%B %Y")}'
        else:
            file_name = f'{main_name} BNNK & BNNP TAHUN {tahun}'
                
        file_path = f'{base_path}/{file_name}.xlsx'
        
        if os.path.exists(base_path): shutil.rmtree(base_path)
        os.makedirs(base_path, exist_ok=True)
        
        base_url = self.request.build_absolute_uri('/')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'{satker.nama_satker}'
        
        # return Response({
        #     'status': True,
        #     'file_name': file_name,
        # })
        
        try:
            # ======= HEADERS =======
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'JUMLAH KEGIATAN',
                'STATUS',
                'NO.',
                'SATUAN KERJA TARGET',
                'TANGGAL',
                'JUMLAH HARI PELAKSANAAN',
                'JUMLAH PESERTA',
                'TUJUAN',
                'HAMBATAN/KENDALA',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI'
            ]
            
            # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

            current_row = 4
            
            # ======= GENERATE HEADERS =======
            for item, header in enumerate(headers[:len(headers)], start=1):
                cell = sheet.cell(row=current_row, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            current_row += 1
            
            no = 0
            no_child = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    no += 1
                
                no_child += 1
                
                formatted_data = {}
                
                # ======= STATUS =======
                status_mapping = {
                    0: 'Belum dikirim',
                    1: 'Dikirim ke BNNP',
                    2: 'Dikirim ke BNN Pusat'
                }
                
                formatted_data['status'] = status_mapping.get(row['status'], '-')
                
                formatted_data['jumlah_kegiatan'] = f"{len(serialized_data)} kali"
                formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])
                formatted_data['jumlah_hari_pelaksanaan'] = f"{row['jumlah_hari_pelaksanaan']} hari"
                formatted_data['jumlah_peserta'] = f"{row['jumlah_peserta']} orang"
                
                for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
                    celcol = sheet.cell(row=current_row, column=col, value=current_group)
                    celcol.font = openpyxl.styles.Font(bold=True)
                    celcol.alignment = Alignment(horizontal='center', vertical='center')
                    
                # ======= COLUMN ADJUSMENT =======
                # Index start dari 1
                
                # No Kegiatan
                col_no = column_index_from_string('E')
                cell_no = sheet.cell(row=current_row, column=col_no, value=current_group)
                cell_no.font = openpyxl.styles.Font(bold=True)
                cell_no.alignment = Alignment(horizontal='center', vertical='center')
                
                # Tanggal
                col_tanggal = column_index_from_string('G')
                cell_tanggal = sheet.cell(row=current_row, column=col_tanggal, value=current_group)
                cell_tanggal.alignment = Alignment(horizontal='center', vertical='center')
                
                # Jumlah Hari Pelaksanaan
                col_jumlah_hari = column_index_from_string('H')
                cell_jumlah_hari = sheet.cell(row=current_row, column=col_jumlah_hari, value=current_group)
                cell_jumlah_hari.alignment = Alignment(horizontal='center', vertical='center')
                
                # Jumlah Peserta
                col_jumlah_peserta = column_index_from_string('I')
                cell_jumlah_peserta = sheet.cell(row=current_row, column=col_jumlah_peserta, value=current_group)
                cell_jumlah_peserta.alignment = Alignment(horizontal='center', vertical='center')
                
                # Dokumentasi
                col_dokumentasi = column_index_from_string('N')
                cell_dokumentasi = sheet.cell(row=current_row, column=14, value=current_group)
                cell_dokumentasi.alignment = Alignment(horizontal='center', vertical='center')

                data_row = [
                    no,
                    row['nama_satker'],
                    formatted_data['jumlah_kegiatan'],
                    formatted_data['status'],
                    no_child,
                    row['nama_satker_target'],
                    formatted_data['tanggal'],
                    formatted_data['jumlah_peserta'],
                    formatted_data['jumlah_hari_pelaksanaan'],
                    row['tujuan'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    # row['dokumentasi'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
                ]

                for item, value in enumerate(data_row[:len(headers)], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1

            # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
            sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
            sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

            # ======= AUTOFIT KOLOM =======
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.3
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
            
            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['N'].width = 35

            # ======= SAVE FILE =======
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diekspor!',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
                
# ======= PEMETAAN POTENSI =======
class DAYATIF_PEMETAAN_POTENSI_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_PEMETAAN_POTENSI.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_PEMETAAN_POTENSI_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.DAYATIF_PEMETAAN_POTENSI_Filters
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_PEMETAAN_POTENSI_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_PEMETAAN_POTENSI, serializers.DAYATIF_PEMETAAN_POTENSI_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_PEMETAAN_POTENSI, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_PEMETAAN_POTENSI)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_PEMETAAN_POTENSI, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_PEMETAAN_POTENSI.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],
                
                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data
                            
    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        satker = self.request.user.profile.satker
        filter_status_pengiriman = request.data.get("status", None)
        filter_rentang_waktu = request.data.get("waktu", None)
        
        serialized_data = self.get_flat_values(request, filter_status_pengiriman, filter_rentang_waktu)
        
        # return Response({
        #     'status': True,
        #     'status': filter_status_pengiriman,
        #     'rentang_waktu': filter_rentang_waktu,
        #     'data': len(serialized_data),
        # })
        
        # ======= GET FILE NAME =======
        file_name = ''
                    
        tahun = datetime.datetime.now().year
        base_path = 'media/kegiatan/pemetaan_potensi/exported'
        main_name = 'PEMETAAN POTENSI SDM DAN SDA KAWASAN RAWAN NARKOBA'
        
        if satker.level != 2:
            if filter_rentang_waktu == 'semua':
                file_name = f'{main_name} {satker.nama_satker.upper()} TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan1':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 1 (Januari - Maret) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan2':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 2 (April - Juni) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan3':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 3 (July - September) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan4':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 4 (Oktober - Desember) TAHUN {tahun}'
            elif filter_rentang_waktu == 'hari_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%d %B %Y")}'
            elif filter_rentang_waktu == 'minggu_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%A, %d %B %Y")}'
            elif filter_rentang_waktu == 'bulan_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%B %Y")}'
        else:
            file_name = f'{main_name} BNNK & BNNP TAHUN {tahun}'
                
        file_path = f'{base_path}/{file_name}.xlsx'
        
        if os.path.exists(base_path): shutil.rmtree(base_path)
        os.makedirs(base_path, exist_ok=True)
        
        base_url = self.request.build_absolute_uri('/')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'{satker.nama_satker}'
        
        # return Response({
        #     'status': True,
        #     'file_name': file_name,
        # })
        
        try:
            # ======= HEADERS =======
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'JUMLAH KEGIATAN',
                'STATUS',
                'NO.',
                # 'SATUAN KERJA TARGET',
                'TANGGAL',
                'LOKASI',
                # 'JUMLAH PESERTA',
                'DESKRIPSI',
                'HAMBATAN/KENDALA',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI'
            ]
            
            # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

            current_row = 4
            
            # ======= GENERATE HEADERS =======
            for item, header in enumerate(headers[:len(headers)], start=1):
                cell = sheet.cell(row=current_row, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            current_row += 1
            
            no = 0
            no_child = 0
            current_group = None
            start_merge_row = 4

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    no += 1
                
                no_child += 1
                
                formatted_data = {}
                
                # ======= STATUS =======
                status_mapping = {
                    0: 'Belum dikirim',
                    1: 'Dikirim ke BNNP',
                    2: 'Dikirim ke BNN Pusat'
                }
                
                formatted_data['status'] = status_mapping.get(row['status'], '-')
                
                formatted_data['jumlah_kegiatan'] = f"{len(serialized_data)} kali"
                formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])
                
                # ======= LOKASI =======
                merged_lokasi = f"{row['nama_desa']}, Kecamatan {row['nama_kecamatan']}, {row['nama_kabupaten']}, Provinsi {row['nama_provinsi']}"
                
                formatted_data['lokasi'] = merged_lokasi
                
                for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
                    celcol = sheet.cell(row=current_row, column=col, value=current_group)
                    celcol.font = openpyxl.styles.Font(bold=True)
                    celcol.alignment = Alignment(horizontal='center', vertical='center')
                    
                # ======= COLUMN ADJUSMENT =======
                # Index start dari 1
                
                # No Kegiatan
                col_no = column_index_from_string('E')
                cell_no = sheet.cell(row=current_row, column=col_no, value=current_group)
                cell_no.font = openpyxl.styles.Font(bold=True)
                cell_no.alignment = Alignment(horizontal='center', vertical='center')
                
                # Tanggal
                col_tanggal = column_index_from_string('F')
                cell_tanggal = sheet.cell(row=current_row, column=col_tanggal, value=current_group)
                cell_tanggal.alignment = Alignment(horizontal='center', vertical='center')
                
                # Dokumentasi
                col_dokumentasi = column_index_from_string('L')
                cell_dokumentasi = sheet.cell(row=current_row, column=col_dokumentasi, value=current_group)
                cell_dokumentasi.alignment = Alignment(horizontal='center', vertical='center')

                data_row = [
                    no,
                    row['nama_satker'],
                    formatted_data['jumlah_kegiatan'],
                    formatted_data['status'],
                    no_child,
                    formatted_data['tanggal'],
                    formatted_data['lokasi'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    # row['dokumentasi'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
                ]

                for item, value in enumerate(data_row[:len(headers)], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1

            # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
            sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
            sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

            # ======= AUTOFIT KOLOM =======
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.3
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
            
            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['L'].width = 35

            # ======= SAVE FILE =======
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diekspor!',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
                    
# ======= PEMETAAN STAKEHOLDER =======
class DAYATIF_PEMETAAN_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_PEMETAAN_STAKEHOLDER.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_PEMETAAN_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = filters.DAYATIF_PEMETAAN_STAKEHOLDER_Filters
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_PEMETAAN_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_PEMETAAN_STAKEHOLDER, serializers.DAYATIF_PEMETAAN_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_PEMETAAN_STAKEHOLDER, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_PEMETAAN_STAKEHOLDER)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_PEMETAAN_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_PEMETAAN_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],
                
                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],
                
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data
                            
    @action(detail=False, methods=['POST'], url_path='export', name='Export Data')
    def export_data(self, request):
        satker = self.request.user.profile.satker
        filter_status_pengiriman = request.data.get("status", None)
        filter_rentang_waktu = request.data.get("waktu", None)
        
        serialized_data = self.get_flat_values(request, filter_status_pengiriman, filter_rentang_waktu)
        
        # return Response({
        #     'status': True,
        #     'status': filter_status_pengiriman,
        #     'rentang_waktu': filter_rentang_waktu,
        #     'data': len(serialized_data),
        # })
        
        # ======= GET FILE NAME =======
        file_name = ''
                    
        tahun = datetime.datetime.now().year
        base_path = 'media/kegiatan/pemetaan_STAKEHOLDER/exported'
        main_name = 'PEMETAAN STAKEHOLDER SDM DAN SDA KAWASAN RAWAN NARKOBA'
        
        if satker.level != 2:
            if filter_rentang_waktu == 'semua':
                file_name = f'{main_name} {satker.nama_satker.upper()} TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan1':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 1 (Januari - Maret) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan2':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 2 (April - Juni) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan3':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 3 (July - September) TAHUN {tahun}'
            elif filter_rentang_waktu == 'triwulan4':
                file_name = f'{main_name} {satker.nama_satker.upper()} TRIWULAN 4 (Oktober - Desember) TAHUN {tahun}'
            elif filter_rentang_waktu == 'hari_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%d %B %Y")}'
            elif filter_rentang_waktu == 'minggu_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%A, %d %B %Y")}'
            elif filter_rentang_waktu == 'bulan_ini':
                file_name = f'{main_name} {satker.nama_satker.upper()} TANGGAL {datetime.datetime.now().strftime("%B %Y")}'
        else:
            file_name = f'{main_name} BNNK & BNNP TAHUN {tahun}'
                
        file_path = f'{base_path}/{file_name}.xlsx'
        
        if os.path.exists(base_path): shutil.rmtree(base_path)
        os.makedirs(base_path, exist_ok=True)
        
        base_url = self.request.build_absolute_uri('/')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'{satker.nama_satker}'
        
        # return Response({
        #     'status': True,
        #     'file_name': file_name,
        # })
        
        try:
            # ======= HEADERS =======
            headers = [
                'NO.',
                'SATUAN KERJA PELAKSANA',
                'JUMLAH KEGIATAN',
                'STATUS',
                'NO.',
                # 'SATUAN KERJA TARGET',
                'TANGGAL',
                'LOKASI',
                # 'JUMLAH PESERTA',
                'DESKRIPSI',
                'HAMBATAN/KENDALA',
                'KESIMPULAN',
                'TINDAK LANJUT',
                'DOKUMENTASI'
            ]
            
            # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

            current_row = 4
            
            # ======= GENERATE HEADERS =======
            for item, header in enumerate(headers[:len(headers)], start=1):
                cell = sheet.cell(row=current_row, column=item, value=header)
                cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            current_row += 1
            
            no = 0
            no_child = 0
            current_group = None

            # ======= MAPPING DATA =======
            for row in serialized_data:
                if row['nama_satker'] != current_group:
                    if current_group is not None:
                        end_merge_row = current_row - 1
                        sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
                        sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
                    current_group = row['nama_satker']
                    start_merge_row = current_row
                    no += 1
                
                no_child += 1
                
                formatted_data = {}
                
                # ======= STATUS =======
                status_mapping = {
                    0: 'Belum dikirim',
                    1: 'Dikirim ke BNNP',
                    2: 'Dikirim ke BNN Pusat'
                }
                
                formatted_data['status'] = status_mapping.get(row['status'], '-')
                
                formatted_data['jumlah_kegiatan'] = f"{len(serialized_data)} kali"
                formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])
                
                # ======= LOKASI =======
                merged_lokasi = f"{row['nama_desa']}, Kecamatan {row['nama_kecamatan']}, {row['nama_kabupaten']}, Provinsi {row['nama_provinsi']}"
                
                formatted_data['lokasi'] = merged_lokasi
                
                for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
                    celcol = sheet.cell(row=current_row, column=col, value=current_group)
                    celcol.font = openpyxl.styles.Font(bold=True)
                    celcol.alignment = Alignment(horizontal='center', vertical='center')
                    
                # ======= COLUMN ADJUSMENT =======
                # Index start dari 1
                
                # No Kegiatan
                col_no = column_index_from_string('E')
                cell_no = sheet.cell(row=current_row, column=col_no, value=current_group)
                cell_no.font = openpyxl.styles.Font(bold=True)
                cell_no.alignment = Alignment(horizontal='center', vertical='center')
                
                # Tanggal
                col_tanggal = column_index_from_string('F')
                cell_tanggal = sheet.cell(row=current_row, column=col_tanggal, value=current_group)
                cell_tanggal.alignment = Alignment(horizontal='center', vertical='center')
                
                # Dokumentasi
                col_dokumentasi = column_index_from_string('L')
                cell_dokumentasi = sheet.cell(row=current_row, column=col_dokumentasi, value=current_group)
                cell_dokumentasi.alignment = Alignment(horizontal='center', vertical='center')

                data_row = [
                    no,
                    row['nama_satker'],
                    formatted_data['jumlah_kegiatan'],
                    formatted_data['status'],
                    no_child,
                    formatted_data['tanggal'],
                    formatted_data['lokasi'],
                    row['deskripsi'],
                    row['kendala'],
                    row['kesimpulan'],
                    row['tindak_lanjut'],
                    # row['dokumentasi'],
                    f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
                ]

                for item, value in enumerate(data_row[:len(headers)], start=1):
                    cell = sheet.cell(row=current_row, column=item, value=value)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )

                current_row += 1

            # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
            end_merge_row = current_row - 1
            sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
            sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
            sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
            sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

            # ======= AUTOFIT KOLOM =======
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.3
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
            
            # ======= ADJUST KOLOM =======
            sheet.column_dimensions['L'].width = 35

            # ======= SAVE FILE =======
            workbook.save(file_path)
            
            return Response({
                'status': True,
                'message': f'Data kegiatan dari {file_name} berhasil diekspor!',
                'file_path': f'/{file_path}'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': False,
                'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
                'error': f'{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # try:
        #     # ======= HEADERS =======
        #     headers = [
        #         'NO.',
        #         'SATUAN KERJA PELAKSANA',
        #         'JUMLAH KEGIATAN',
        #         'STATUS',
        #         'NO.',
        #         # 'SATUAN KERJA TARGET',
        #         'TANGGAL',
        #         'LOKASI',
        #         'STAKEHOLDERS',
        #         'DESKRIPSI',
        #         'HAMBATAN/KENDALA',
        #         'KESIMPULAN',
        #         'TINDAK LANJUT',
        #         'DOKUMENTASI'
        #     ]
            
        #     # TUJUAN [J] SAMPAI KESIMPULAN [L] Ada Parent colspan 3 nya

        #     current_row = 4
            
        #     # ======= GENERATE HEADERS =======
        #     for item, header in enumerate(headers[:len(headers)], start=1):
        #         cell = sheet.cell(row=current_row, column=item, value=header)
        #         cell.fill = openpyxl.styles.PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        #         cell.font = openpyxl.styles.Font(bold=True)
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
        #     current_row += 1
            
        #     no = 0
        #     no_child = 0
        #     current_group = None

        #     # ======= MAPPING DATA =======
        #     for row in serialized_data:
        #         if row['nama_satker'] != current_group:
        #             if current_group is not None:
        #                 end_merge_row = current_row - 1
        #                 sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
        #                 sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')  # Gabungkan sel grup
        #             current_group = row['nama_satker']
        #             start_merge_row = current_row
        #             no += 1
                
        #         no_child += 1
                
        #         formatted_data = {}
                
        #         # ======= STATUS =======
        #         status_mapping = {
        #             0: 'Belum dikirim',
        #             1: 'Dikirim ke BNNP',
        #             2: 'Dikirim ke BNN Pusat'
        #         }
                
        #         formatted_data['status'] = status_mapping.get(row['status'], '-')
                
        #         formatted_data['jumlah_kegiatan'] = f"{len(serialized_data)} kali"
        #         formatted_data['tanggal'] = get_tanggal_kegiatan(row['tanggal_awal'], row['tanggal_akhir'])
        #         formatted_data['stakeholders'] = ', '.join([stakeholder['nama'] for stakeholder in row['stakeholders']])
                
        #         # ======= LOKASI =======
        #         merged_lokasi = f"{row['nama_desa']}, Kecamatan {row['nama_kecamatan']}, {row['nama_kabupaten']}, Provinsi {row['nama_provinsi']}"
                
        #         formatted_data['lokasi'] = merged_lokasi
                
        #         for col in range(1, 5):  # Menggabungkan kolom dari A hingga D
        #             celcol = sheet.cell(row=current_row, column=col, value=current_group)
        #             celcol.font = openpyxl.styles.Font(bold=True)
        #             celcol.alignment = Alignment(horizontal='center', vertical='center')
                    
        #         # ======= COLUMN ADJUSMENT =======
        #         # Index start dari 1
                
        #         # No Kegiatan
        #         col_no = column_index_from_string('E')
        #         cell_no = sheet.cell(row=current_row, column=col_no, value=current_group)
        #         cell_no.font = openpyxl.styles.Font(bold=True)
        #         cell_no.alignment = Alignment(horizontal='center', vertical='center')
                
        #         # Tanggal
        #         col_tanggal = column_index_from_string('F')
        #         cell_tanggal = sheet.cell(row=current_row, column=col_tanggal, value=current_group)
        #         cell_tanggal.alignment = Alignment(horizontal='center', vertical='center')
                
        #         # Dokumentasi
        #         col_dokumentasi = column_index_from_string('M')
        #         cell_dokumentasi = sheet.cell(row=current_row, column=col_dokumentasi, value=current_group)
        #         cell_dokumentasi.alignment = Alignment(horizontal='center', vertical='center')

        #         data_row = [
        #             no,
        #             row['nama_satker'],
        #             formatted_data['jumlah_kegiatan'],
        #             formatted_data['status'],
        #             no_child,
        #             formatted_data['tanggal'],
        #             formatted_data['lokasi'],
        #             formatted_data['stakeholders'],
        #             row['deskripsi'],
        #             row['kendala'],
        #             row['kesimpulan'],
        #             row['tindak_lanjut'],
        #             # row['dokumentasi'],
        #             f'=HYPERLINK("{base_url + row["dokumentasi"]}","Dokumentasi")',
        #         ]

        #         for item, value in enumerate(data_row[:len(headers)], start=1):
        #             cell = sheet.cell(row=current_row, column=item, value=value)
        #             cell.border = openpyxl.styles.Border(
        #                 left=openpyxl.styles.Side(style='thin'),
        #                 right=openpyxl.styles.Side(style='thin'),
        #                 top=openpyxl.styles.Side(style='thin'),
        #                 bottom=openpyxl.styles.Side(style='thin')
        #             )

        #         current_row += 1

        #     # ======= MERGE CELL UNTUK GRUP TERAKHIR =======
        #     end_merge_row = current_row - 1
        #     sheet.merge_cells(f'A{start_merge_row}:A{end_merge_row}')
        #     sheet.merge_cells(f'B{start_merge_row}:B{end_merge_row}')
        #     sheet.merge_cells(f'C{start_merge_row}:C{end_merge_row}')
        #     sheet.merge_cells(f'D{start_merge_row}:D{end_merge_row}')

        #     # ======= AUTOFIT KOLOM =======
        #     for column in sheet.columns:
        #         max_length = 0
        #         for cell in column:
        #             if cell.value:
        #                 max_length = max(max_length, len(str(cell.value)))
        #         adjusted_width = (max_length + 2) * 1.3
        #         sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
            
        #     # ======= ADJUST KOLOM =======
        #     sheet.column_dimensions['M'].width = 35

        #     # ======= SAVE FILE =======
        #     workbook.save(file_path)
            
        #     return Response({
        #         'status': True,
        #         'message': f'Data kegiatan dari {file_name} berhasil diekspor!',
        #         'file_path': f'/{file_path}'
        #     }, status=status.HTTP_200_OK)
        # except Exception as e:
        #     return Response({
        #         'status': False,
        #         'message': f'Gagal mengekspor daftar kegiatan dari Satuan Kerja {satker.nama_satker}',
        #         'error': f'{str(e)}'
        #     }, status=status.HTTP_400_BAD_REQUEST)

# ======= RAPAT SINERGI STAKEHOLDER =======
class DAYATIF_RAPAT_SINERGI_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_RAPAT_SINERGI_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_RAPAT_SINERGI_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER, serializers.DAYATIF_RAPAT_SINERGI_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_RAPAT_SINERGI_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],
                
                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],
                
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data

# ======= BIMBINGAN_TEKNIS_STAKEHOLDER =======
class DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER, serializers.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_BIMBINGAN_TEKNIS_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],
                
                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],
                
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data

# ======= BIMBINGAN_TEKNIS_LIFESKILL =======
class DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.all().order_by('-tanggal_awal')
    serializer_class = serializers.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_LIST_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL, serializers.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_BIMBINGAN_TEKNIS_LIFESKILL.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'tanggal_awal', 'tanggal_akhir',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'deskripsi', 'kendala', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'tanggal_awal': item['tanggal_awal'],
                'tanggal_akhir': item['tanggal_akhir'],
                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],
                
                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'stakeholders': item['stakeholders'],
                
                'deskripsi': item['deskripsi'],
                'kendala': item['kendala'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data

# ======= DUKUNGAN_STAKEHOLDER =======
class DAYATIF_DUKUNGAN_STAKEHOLDER_ViewSet(viewsets.ModelViewSet):
    queryset = models.DAYATIF_DUKUNGAN_STAKEHOLDER.objects.all()
    serializer_class = serializers.DAYATIF_DUKUNGAN_STAKEHOLDER_Serializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = pagination.Page10NumberPagination
    filter_backends = [DjangoFilterBackend]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=set_created_kegiatan_status(self.request), satker=self.request.user.profile.satker)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        
    @action(detail=False, methods=['GET'], url_path='list', name='List Data')
    def data_list(self, request):
        return get_data_list(request, self.queryset, serializers.DAYATIF_DUKUNGAN_STAKEHOLDER_Serializer)

    @action(detail=False, methods=['GET'], url_path='list/bnnk', name='List Data BNNK')
    def data_list_bnnk(self, request):
        return get_data_list_bnnk(self.request.user.profile.satker, models.DAYATIF_DUKUNGAN_STAKEHOLDER, serializers.DAYATIF_DUKUNGAN_STAKEHOLDER_Serializer)

    @action(detail=True, methods=['DELETE'], url_path='delete_all_kegiatan', name='Hapus semua Kegiatan')
    def delete_all_kegiatan(self, request, pk=None):
        return delete_all_kegiatan_helper(models.DAYATIF_DUKUNGAN_STAKEHOLDER, pk)
    
    @action(detail=False, methods=['POST'], url_path='semua_kegiatan', name='Aksi untuk semua Kegiatan')
    def semua_kegiatan(self, request):
        return aksi_semua_kegiatan(request, models.DAYATIF_DUKUNGAN_STAKEHOLDER)
    
    @action(detail=False, methods=['POST'], url_path='kirim_kegiatan', name='Kirim Kegiatan')
    def kirim_kegiatan(self, request):
        return kirim_kegiatan_helper(models.DAYATIF_DUKUNGAN_STAKEHOLDER, request)

    def get_flat_values(self, request, status, waktu):
        satker = self.request.user.profile.satker

        data = models.DAYATIF_DUKUNGAN_STAKEHOLDER.objects.values(
            'id', 'satker_id', 'satker__nama_satker', 'stakeholder', 'jumlah_peserta', 'jenis', 'bentuk', 'jumlah',
            'desa', 'kecamatan', 'kabupaten', 'provinsi', 'nama_desa', 'nama_kecamatan', 'nama_kabupaten', 'nama_provinsi', 'jumlah_sasaran', 'pengaruh', 'kesimpulan', 'tindak_lanjut', 'dokumentasi', 'status'
        ).order_by('satker__nama_satker')

        data = get_data_list_queryset(request, data)
        
        print('Panjang data sebelum filter:', len(data))
        
        data = get_filtered_data(satker, data, status, waktu)
        
        print('Panjang data sesudah filter:', len(data))
    
        serialized_data = []
        for item in data:
            serialized_item = {
                'id': item['id'],
                'satker_id': item['satker_id'],
                'nama_satker': item['satker__nama_satker'],
                'stakeholder': item['stakeholder'],
                'jumlah_peserta': item['jumlah_peserta'],
                'jenis': item['jenis'],
                'bentuk': item['bentuk'],
                'jumlah': item['jumlah'],

                'desa': item['desa'],
                'kecamatan': item['kecamatan'],
                'kabupaten': item['kabupaten'],
                'provinsi': item['provinsi'],
                
                'nama_desa': item['nama_desa'],
                'nama_kecamatan': item['nama_kecamatan'],
                'nama_kabupaten': item['nama_kabupaten'],
                'nama_provinsi': item['nama_provinsi'],
                'jumlah_sasaran': item['jumlah_sasaran'],
                
                'pengaruh': item['pengaruh'],
                'kesimpulan': item['kesimpulan'],
                'tindak_lanjut': item['tindak_lanjut'],
                'dokumentasi': item['dokumentasi'],
                'status': item['status'],
                'satker_level': satker.level
            }
            serialized_data.append(serialized_item)
        
        return serialized_data
